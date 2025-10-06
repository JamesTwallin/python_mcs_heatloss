"""Analyze the MCS Heat Pump Calculator Excel file structure."""
import sys
import openpyxl
from openpyxl import load_workbook

# Ensure UTF-8 encoding for output
sys.stdout.reconfigure(encoding='utf-8')

def analyze_excel_structure(filepath):
    """Analyze and print the structure of the Excel file."""
    wb = load_workbook(filepath, data_only=False, keep_vba=True)

    print(f"Workbook contains {len(wb.sheetnames)} sheets:")
    print("-" * 80)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"  Dimensions: {ws.dimensions}")
        print(f"  Max row: {ws.max_row}, Max col: {ws.max_column}")

        # Find cells with values or formulas
        cells_with_content = []
        for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row)):
            for cell in row:
                if cell.value is not None:
                    cells_with_content.append({
                        'cell': cell.coordinate,
                        'value': str(cell.value)[:100],  # Truncate long values
                        'is_formula': str(cell.value).startswith('=') if isinstance(cell.value, str) else False
                    })

        if cells_with_content:
            print(f"  Sample content (first 50 rows):")
            for item in cells_with_content[:20]:  # Show first 20 items
                formula_marker = " [FORMULA]" if item['is_formula'] else ""
                # Clean value to remove problematic characters
                clean_value = item['value'].encode('ascii', 'ignore').decode('ascii')
                print(f"    {item['cell']}: {clean_value}{formula_marker}")

    wb.close()

if __name__ == "__main__":
    filepath = r"C:\Users\User\Documents\GitHub\python_mcs_heatloss\MCS-Heat-Pump-Calculator-Version-1.10-unlocked-1.xlsm"
    analyze_excel_structure(filepath)
