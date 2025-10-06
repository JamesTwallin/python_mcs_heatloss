"""Extract a complete test case from Excel file to validate Python implementation."""
import openpyxl
from openpyxl import load_workbook
import json


def extract_room_data_from_excel(filepath, sheet_name='1'):
    """Extract all room data from a specific room sheet."""
    wb = load_workbook(filepath, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found")
        return None

    ws = wb[sheet_name]

    # Extract room data
    room_data = {
        'sheet_name': sheet_name,
        'room_info': {},
        'fabric_elements': [],
        'ventilation': {},
        'heat_loss': {}
    }

    # Room information (rows 1-10 typically)
    print(f"\nAnalyzing Sheet: {sheet_name}")
    print("=" * 80)

    # Extract key cells - adjust based on actual Excel structure
    for row in range(1, 65):
        for col in range(2, 26):  # B to Y columns
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                coord = cell.coordinate
                value = cell.value
                # Print to help identify structure
                if isinstance(value, (int, float)) or (isinstance(value, str) and len(str(value)) < 50):
                    print(f"{coord}: {value}")

    wb.close()
    return room_data


def extract_design_details(filepath):
    """Extract design details and summary from Excel."""
    wb = load_workbook(filepath, data_only=True)

    if 'Design Details' not in wb.sheetnames:
        print("Design Details sheet not found")
        wb.close()
        return None

    ws = wb['Design Details']

    design_data = {}

    print("\nDesign Details Sheet - Key Values:")
    print("=" * 80)

    # Scan specific areas known to contain important data
    areas = [
        (2, 2, 30, 30),    # Top left area - project info
        (26, 2, 30, 10),   # Summary area
        (87, 2, 117, 10),  # Room list area
    ]

    for row_start, col_start, row_end, col_end in areas:
        for row in range(row_start, row_end):
            for col in range(col_start, col_end):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    coord = cell.coordinate
                    value = cell.value
                    if isinstance(value, (int, float)) or (isinstance(value, str) and len(str(value)) < 80):
                        print(f"{coord}: {value}")

    wb.close()
    return design_data


def main():
    """Extract test case from Excel."""
    filepath = r"C:\Users\User\Documents\GitHub\python_mcs_heatloss\MCS-Heat-Pump-Calculator-Version-1.10-unlocked-1.xlsm"

    print("Extracting Design Details...")
    design_data = extract_design_details(filepath)

    print("\n" + "=" * 80)
    print("Extracting Room 1 Data...")
    room1_data = extract_room_data_from_excel(filepath, '1')

    # Save to file for analysis
    output = {
        'design_details': design_data,
        'room_1': room1_data
    }

    with open('excel_extracted_data.json', 'w') as f:
        json.dump(output, f, indent=2, default=str)

    print("\n" + "=" * 80)
    print("Data saved to: excel_extracted_data.json")
    print("\nNext steps:")
    print("1. Review the extracted data to identify the exact cell locations")
    print("2. Create a test case that populates Python with the same inputs")
    print("3. Compare outputs to ensure they match")


if __name__ == '__main__':
    main()
