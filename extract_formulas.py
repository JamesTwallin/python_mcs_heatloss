"""Extract key formulas and data from the MCS Heat Pump Calculator."""
import openpyxl
from openpyxl import load_workbook
import json

def extract_key_data(filepath):
    """Extract key calculation data and formulas from Excel file."""
    wb = load_workbook(filepath, data_only=False)

    results = {}

    # Extract Design Details sheet - this is the main input/output sheet
    if 'Design Details' in wb.sheetnames:
        ws = wb['Design Details']
        design_data = {}
        for row in ws.iter_rows(min_row=1, max_row=120, values_only=False):
            for cell in row:
                if cell.value is not None:
                    coord = cell.coordinate
                    value = cell.value
                    design_data[coord] = {
                        'value': str(value)[:200],
                        'is_formula': str(value).startswith('=') if isinstance(value, str) else False
                    }
        results['Design Details'] = design_data

    # Extract degree days lookup data
    if 'Post Code Degree Days' in wb.sheetnames:
        ws = wb['Post Code Degree Days']
        degree_days = {}
        for row in ws.iter_rows(min_row=3, max_row=126, values_only=True):
            if row[0]:  # Postcode area
                degree_days[row[0]] = {
                    'temp': row[1],
                    'degree_days': row[2],
                    'location': row[6] if len(row) > 6 else None
                }
        results['degree_days'] = degree_days

    # Extract U-value tables
    if 'floor u values' in wb.sheetnames:
        ws = wb['floor u values']
        floor_data = []
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
            row_data = []
            for cell in row:
                if cell.value:
                    row_data.append({
                        'coord': cell.coordinate,
                        'value': str(cell.value)[:100],
                        'is_formula': str(cell.value).startswith('=') if isinstance(cell.value, str) else False
                    })
            if row_data:
                floor_data.append(row_data)
        results['floor_u_values'] = floor_data

    # Extract Room 1 structure as template
    if '1' in wb.sheetnames:
        ws = wb['1']
        room_structure = {}
        for row in ws.iter_rows(min_row=1, max_row=61, values_only=False):
            for cell in row:
                if cell.value is not None:
                    coord = cell.coordinate
                    room_structure[coord] = {
                        'value': str(cell.value)[:150],
                        'is_formula': str(cell.value).startswith('=') if isinstance(cell.value, str) else False
                    }
        results['room_template'] = room_structure

    wb.close()

    # Save to JSON for analysis
    with open('calculator_structure.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)

    print(f"Extracted data from {len(results)} sections")
    print(f"Degree days for {len(results.get('degree_days', {}))} postcode areas")
    print(f"Design Details contains {len(results.get('Design Details', {}))} cells")
    print(f"Room template contains {len(results.get('room_template', {}))} cells")
    print("\nSaved to calculator_structure.json")

if __name__ == "__main__":
    filepath = r"C:\Users\User\Documents\GitHub\python_mcs_heatloss\MCS-Heat-Pump-Calculator-Version-1.10-unlocked-1.xlsm"
    extract_key_data(filepath)
